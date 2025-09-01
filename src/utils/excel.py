from __future__ import annotations
from io import BytesIO
from typing import Dict, Any, List
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def build_order_xlsx(order: Dict[str, Any], user: Dict[str, Any] | None = None) -> bytes:
    """
    Генерирует XLSX по заказу.
    order: документ из Mongo c полем items=[...]
    user: словарь (опционально) с данными профиля
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Order"

    # Шапка
    ws.append(["Order ID", str(order.get("_id", ""))])
    ws.append(["User ID", str(order.get("user_id", ""))])
    if user:
        ws.append(["Username", user.get("username") or ""])
        ws.append(["Full name", user.get("full_name") or ""])
        ws.append(["Phone", user.get("phone") or ""])
        ws.append(["Delivery", user.get("delivery") or ""])
    ws.append([])

    # Заголовки таблицы
    headers = ["#", "PhotoFileId", "Link", "Size", "Color", "Price CNY", "Qty"]
    ws.append(headers)

    items: List[Dict[str, Any]] = order.get("items", [])
    for i, it in enumerate(items, start=1):
        ws.append([
            i,
            it.get("photo_file_id", "") or "",
            it.get("link", "") or "",
            it.get("size", "") or "",
            it.get("color", "") or "",
            it.get("price_cny", "") if it.get("price_cny") is not None else "",
            it.get("qty", "") if it.get("qty") is not None else "",
        ])

    # Небольшая авто-ширина
    for col_idx, _ in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 18

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
